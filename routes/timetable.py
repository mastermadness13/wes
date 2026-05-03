from datetime import datetime
from io import BytesIO

from flask import Blueprint, flash, g, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import db
from routes import (
    courses_timetable_admin_required,
    current_department_name,
    current_department_id,
    current_role,
    current_user_id,
    login_required,
    super_admin_required,
)
from routes.access import admin_department_required, department_name_allowed, is_super_admin


timetable_bp = Blueprint('timetable', __name__)

DAY_NAMES = ['الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
DAY_LABELS = {day: day for day in DAY_NAMES}


def _safe_int(value, default=None):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_time(value):
    if not value:
        return None
    return datetime.strptime(value, '%H:%M').time()


def _minutes(time_value):
    return time_value.hour * 60 + time_value.minute


def _time_ranges_overlap(start_a, end_a, start_b, end_b):
    if not all([start_a, end_a, start_b, end_b]):
        return False
    return _minutes(start_a) < _minutes(end_b) and _minutes(start_b) < _minutes(end_a)


def _get_periods(conn, enabled_only=False):
    if 'periods' not in g:
        sql = 'SELECT code, label, start_time, end_time, is_enabled, sort_order FROM period_settings ORDER BY sort_order, code'
        g.periods = []
        for row in conn.execute(sql).fetchall():
            g.periods.append(
                {
                    'code': row['code'],
                    'label': row['label'],
                    'start_time': row['start_time'],
                    'end_time': row['end_time'],
                    'is_enabled': bool(row['is_enabled']),
                    'sort_order': row['sort_order'],
                    'start_obj': _parse_time(row['start_time']),
                    'end_obj': _parse_time(row['end_time']),
                }
            )
    if enabled_only:
        return [p for p in g.periods if p['is_enabled']]
    return g.periods


def _period_map(conn):
    if 'period_map' not in g:
        g.period_map = {p['code']: p for p in _get_periods(conn)}
    return g.period_map


def _validate_period_definitions(periods):
    enabled = [period for period in periods if period['is_enabled']]
    if not enabled:
        return 'At least one period must be enabled.'

    enabled.sort(key=lambda item: item['sort_order'])
    if enabled[0]['start_time'] != '09:00':
        return 'The first enabled period must start at 09:00.'

    for period in enabled:
        if _minutes(period['start_obj']) >= _minutes(period['end_obj']):
            return f"{period['label']} must end after it starts."

    for index, current in enumerate(enabled):
        for other in enabled[index + 1:]:
            if _time_ranges_overlap(current['start_obj'], current['end_obj'], other['start_obj'], other['end_obj']):
                return f"{current['label']} overlaps with {other['label']}."

    return None


def _periods_from_form(form):
    periods = []
    for code in ['A', 'B', 'C']:
        start_value = (form.get(f'{code}_start') or '').strip()
        end_value = (form.get(f'{code}_end') or '').strip()
        label = (form.get(f'{code}_label') or f'Period {code}').strip()
        enabled = 1 if form.get(f'{code}_enabled') == '1' else 0
        try:
            start_obj = _parse_time(start_value)
            end_obj = _parse_time(end_value)
        except ValueError:
            return None, 'Time must use HH:MM format.'
        periods.append(
            {
                'code': code,
                'label': label,
                'start_time': start_value,
                'end_time': end_value,
                'is_enabled': enabled,
                'sort_order': {'A': 1, 'B': 2, 'C': 3}[code],
                'start_obj': start_obj,
                'end_obj': end_obj,
            }
        )
    return periods, None


def _build_allowed_semesters(selected_department_name, department_semester_map):
    fallback = [1, 2, 3, 4, 5, 6, 7]
    if selected_department_name and selected_department_name in department_semester_map:
        try:
            max_sem = max(1, int(department_semester_map[selected_department_name]))
        except Exception:
            max_sem = 1
        return list(range(1, max_sem + 1))

    if department_semester_map:
        values = set()
        for max_sem_raw in department_semester_map.values():
            try:
                max_sem = max(1, int(max_sem_raw))
            except Exception:
                continue
            values.update(range(1, max_sem + 1))
        if values:
            return sorted(values)

    return fallback


def _fetch_departments(conn):
    if 'departments' not in g:
        rows = conn.execute('SELECT id, name, semesters FROM departments ORDER BY name').fetchall()
        g.departments = [dict(row) for row in rows]
    return g.departments


def _department_id_map(conn):
    if 'dept_id_map' not in g:
        g.dept_id_map = {d['id']: d for d in _fetch_departments(conn)}
    return g.dept_id_map


def _scoped_department_name(conn):
    department_name = current_department_name(conn)
    if department_name:
        return department_name
    if current_role() == 'admin':
        fallback_label = (session.get('label') or '').strip()
        if fallback_label:
            return fallback_label
    return None


def _selected_department_id():
    return _safe_int(request.values.get('department_id'))


def _selected_department_name(conn, fallback=None):
    if not is_super_admin():
        return _scoped_department_name(conn)

    department_id = _selected_department_id()
    if department_id is not None:
        dept = _department_id_map(conn).get(department_id)
        if dept:
            return dept['name']

    return fallback


def _fetch_scoped_resources(conn, table_name, include_all=False, owner_user_id=None, department_name=None):
    """Unified helper to fetch courses or teachers with correct scoping."""
    params = []
    department_name = (department_name or '').strip()
    is_courses = (table_name == 'courses')
    order_by = 'department, year, name' if is_courses else 'department, name'
    # Teachers allow NULL/empty departments for "general" teachers
    dept_clause = "LOWER(TRIM(department)) = LOWER(TRIM(?))" if is_courses else "(LOWER(TRIM(department)) = LOWER(TRIM(?)) OR department IS NULL OR TRIM(department) = '')"

    if include_all:
        sql = f"SELECT t.*, u.username AS owner_username, u.label AS owner_label FROM {table_name} t LEFT JOIN users u ON u.id = t.user_id"
        if department_name:
            sql += f" WHERE {dept_clause.replace('department', 't.department')}"
            params.append(department_name)
    else:
        sql = f"SELECT * FROM {table_name}"
        if department_name:
            sql += f" WHERE {dept_clause}"
            params.append(department_name)
        else:
            sql += " WHERE user_id = ?"
            params.append(owner_user_id)

    sql += f" ORDER BY {order_by}"
    return [dict(row) for row in conn.execute(sql, params).fetchall()]


def _fetch_courses(conn, include_all=False, owner_user_id=None, department_name=None):
    return _fetch_scoped_resources(conn, 'courses', include_all=include_all, owner_user_id=owner_user_id, department_name=department_name)


def _fetch_teachers(conn, owner_user_id=None, include_all=False, department_name=None):
    return _fetch_scoped_resources(conn, 'teachers', include_all=include_all, owner_user_id=owner_user_id, department_name=department_name)


def _fetch_rooms(conn):
    rows = conn.execute('SELECT * FROM rooms ORDER BY name').fetchall()
    return [dict(row) for row in rows]


def _entry_with_relations(conn, entry_id):
    row = conn.execute(
        """
        SELECT t.*, c.name AS course_name, c.department AS course_department,
               te.name AS teacher_name,
               r.name AS room_name, r.name_ar AS room_name_ar,
               u.username AS owner_username, u.label AS owner_label,
               d.id AS department_id, d.name AS department_name
        FROM timetable t
        JOIN courses c ON c.id = t.course_id
        JOIN teachers te ON te.id = t.teacher_id
        JOIN rooms r ON r.id = t.room_id
        LEFT JOIN users u ON u.id = t.user_id
        LEFT JOIN departments d ON d.id = t.department_id
        WHERE t.id = ?
        """,
        (entry_id,),
    ).fetchone()
    return dict(row) if row else None


def _can_access_entry(conn, entry):
    if not entry:
        return False
    if is_super_admin() or entry['user_id'] == current_user_id():
        return True

    # Optimization: Use pre-fetched department from the join if available to avoid N+1 queries per cell
    dept = entry.get('course_department')
    if not dept:
        course = conn.execute('SELECT department FROM courses WHERE id = ?', (entry['course_id'],)).fetchone()
        dept = course['department'] if course else None

    return bool(dept and department_name_allowed(dept, conn))


def _can_delete_critical():
    return current_role() == 'super_admin'


def _history_actor():
    return current_user_id(), session.get('username') or session.get('label') or 'System'


def _timetable_history_message(action, entry):
    actor_username = session.get('username') or 'System'
    course_name = entry.get('course_name') or f"Course #{entry.get('course_id')}"
    teacher_name = entry.get('teacher_name') or f"Teacher #{entry.get('teacher_id')}"
    room_name = entry.get('room_name_ar') or entry.get('room_name') or f"Room #{entry.get('room_id')}"
    day = entry.get('day') or ''
    semester = entry.get('semester') or ''
    section = entry.get('section') or ''
    action_word = {'ADD': 'added', 'EDIT': 'updated', 'DELETE': 'deleted'}.get(action, 'changed')
    return f"{actor_username} {action_word} timetable entry for {course_name} on {day}, semester {semester}, period {section}, room {room_name}, teacher {teacher_name}."


def _is_resource_accessible(conn, table_name, resource_id, owner_user_id):
    """Unified check to see if a specific resource (course, teacher, room) is accessible."""
    if resource_id is None:
        return False

    row = conn.execute(f"SELECT * FROM {table_name} WHERE id = ?", (resource_id,)).fetchone()
    if not row:
        return False

    # convert to dict for safe .get() usage
    r = dict(row)

    if is_super_admin():
        return True

    if table_name == 'rooms':
        # Currently, rooms are generally accessible if they exist.
        return True

    # Teachers and Courses use 'department' name and 'user_id' for ownership/scoping
    dept_name = (r.get('department') or '').strip()

    # Special rule: Teachers with no department are "general" and accessible to everyone
    if table_name == 'teachers' and not dept_name:
        return True

    return r.get('user_id') == owner_user_id or department_name_allowed(dept_name, conn)


def _validate_owner_resources(conn, owner_user_id, course_id, teacher_id, room_id):
    if not _is_resource_accessible(conn, 'courses', course_id, owner_user_id):
        return 'المقرر المختار غير متاح.'
    if not _is_resource_accessible(conn, 'teachers', teacher_id, owner_user_id):
        return 'المدرس المختار غير متاح.'
    if not _is_resource_accessible(conn, 'rooms', room_id, owner_user_id):
        return 'القاعة المختارة غير موجودة.'
    return None


def _resolve_time_range(period_lookup, period_code, start_time=None, end_time=None):
    period = period_lookup.get(period_code)
    if not period:
        return None, None
    start_obj = period['start_obj']
    end_obj = period['end_obj']
    if start_time and end_time:
        try:
            start_obj = _parse_time(start_time)
            end_obj = _parse_time(end_time)
        except ValueError:
            return None, None
    return start_obj, end_obj


def _validate_schedule_conflicts(
    conn,
    *,
    day,
    semester,
    period_code,
    teacher_id,
    department_id,
    room_id,
    exclude_entry_id=None,
    start_time=None,
    end_time=None,
):
    period_lookup = _period_map(conn)
    target_period = period_lookup.get(period_code)
    if not target_period or not target_period['is_enabled']:
        return 'الفترة المختارة غير مفعلة.'

    target_start, target_end = _resolve_time_range(period_lookup, period_code, start_time, end_time)
    if not target_start or not target_end or _minutes(target_start) >= _minutes(target_end):
        return 'النطاق الزمني المختار غير صالح.'

    # 1. Structural conflict: One slot per department per semester
    sql_dept = """
        SELECT t.id, c.name AS course_name 
        FROM timetable t 
        JOIN courses c ON c.id = t.course_id
        WHERE t.department_id = ? 
        AND t.day = ? AND t.semester = ? AND t.section = ?
    """
    params_dept = [department_id, day, semester, period_code]
    if exclude_entry_id is not None:
        sql_dept += ' AND t.id != ?'
        params_dept.append(exclude_entry_id)
    
    conflict_dept = conn.execute(sql_dept, params_dept).fetchone()
    if conflict_dept:
        return f"هناك حصة مسجلة مسبقاً لهذا القسم (مقرر: {conflict_dept['course_name']}) في هذا اليوم والفترة."

    # 2. Resource conflicts: Teacher/Room overlaps (global)
    sql = """
        SELECT t.id, t.section, t.teacher_id, t.room_id, t.start_time, t.end_time,
               c.name AS course_name, te.name AS teacher_name, 
               r.name_ar AS room_name, r.name AS room_name_en
        FROM timetable t
        JOIN courses c ON c.id = t.course_id
        JOIN teachers te ON te.id = t.teacher_id
        JOIN rooms r ON r.id = t.room_id
        WHERE t.day = ? AND t.semester = ?
    """
    params = [day, semester]
    if exclude_entry_id is not None:
        sql += ' AND t.id != ?'
        params.append(exclude_entry_id)

    for row in conn.execute(sql, params).fetchall():
        existing_start, existing_end = _resolve_time_range(period_lookup, row['section'], row['start_time'], row['end_time'])
        if not _time_ranges_overlap(target_start, target_end, existing_start, existing_end):
            continue
            
        room_display = row['room_name'] or row['room_name_en']
        if row['room_id'] == room_id:
            return f"القاعة ({room_display}) محجوزة بالفعل لمقرر '{row['course_name']}' مع المدرس {row['teacher_name']} في نفس الوقت."
        if row['teacher_id'] == teacher_id:
            return f"المدرس {row['teacher_name']} منشغل حالياً في قاعة ({room_display}) لتدريس مقرر '{row['course_name']}'."

    return None


def check_timetable_conflict(conn, department_id, day, semester, section, exclude_id=None):
    """Return a conflicting timetable row for the given department_id/day/semester/section, or None."""
    # department_id may be an integer id OR a department name (string).
    params = []
    base_sql = """
        SELECT t.id, c.name AS course_name
        FROM timetable t
        JOIN courses c ON c.id = t.course_id
        WHERE {dept_clause}
          AND t.day = ? AND t.semester = ? AND t.section = ?
    """

    # Normalize inputs
    dept_raw = department_id
    dept_clause = ''
    if dept_raw is None:
        return None

    try:
        dept_int = int(dept_raw)
        # match by id OR by course.department name
        dept_clause = '(t.department_id = ? OR LOWER(TRIM(c.department)) = LOWER(TRIM(?)))'
        params.extend([dept_int, str(dept_raw)])
    except Exception:
        # treat as department name
        dept_clause = 'LOWER(TRIM(c.department)) = LOWER(TRIM(?))'
        params.append(str(dept_raw))

    sql = base_sql.format(dept_clause=dept_clause)
    params.extend([day, semester, section])
    if exclude_id is not None:
        sql += ' AND t.id != ?'
        params.append(exclude_id)
    return conn.execute(sql, params).fetchone()


def _process_timetable_save(conn, entry_id, data):
    """Common logic for creating or updating a timetable entry."""
    day = (data.get('day') or '').strip()
    semester = _safe_int(data.get('semester'))
    period_code = (data.get('section') or data.get('period_code') or '').strip()
    course_id = _safe_int(data.get('course_id'))
    teacher_id = _safe_int(data.get('teacher_id'))
    room_id = _safe_int(data.get('room_id'))
    # department_id must be provided by the form; fallback to course's department id or current user's department
    department_id = _safe_int(data.get('department_id'))
    owner_user_id = current_user_id()

    try:
        # BEGIN IMMEDIATE to prevent race conditions during write operations
        conn.execute("BEGIN IMMEDIATE")

        # 1. Resource existence & metadata validation (Fetch ONCE)
        course_row = conn.execute('SELECT id, department, name, user_id FROM courses WHERE id = ?', (course_id,)).fetchone()
        teacher_row = conn.execute('SELECT id, name, department, user_id FROM teachers WHERE id = ?', (teacher_id,)).fetchone()
        room_row = conn.execute('SELECT id, name FROM rooms WHERE id = ?', (room_id,)).fetchone()

        # normalize sqlite3.Row to dict for safe attribute access using .get()
        if course_row:
            course_row = dict(course_row)
        if teacher_row:
            teacher_row = dict(teacher_row)
        if room_row:
            room_row = dict(room_row)

        if not course_row:
            conn.rollback()
            return None, 'المقرر المختار غير موجود.'
        if teacher_id and not teacher_row:
            conn.rollback()
            return None, 'المدرس المختار غير موجود.'
        if room_id and not room_row:
            conn.rollback()
            return None, 'القاعة المختارة غير موجودة.'

        # 2. Accessibility Checks (Inline to avoid duplicate SQL)
        if not is_super_admin():
            if not department_name_allowed(course_row['department'], conn):
                conn.rollback()
                return None, 'المقرر المختار غير متاح لقسمك.'
            t_dept = (teacher_row['department'] or '').strip() if teacher_row else ''
            if t_dept and not department_name_allowed(t_dept, conn):
                conn.rollback()
                return None, 'المدرس المختار غير متاح لقسمك.'

        if not all([day, semester is not None, period_code, course_id, teacher_id, room_id]):
            conn.rollback()
            return None, 'جميع الحقول مطلوبة.'

        # Resolve missing department_id from course's department name if not provided
        if not department_id:
            course_dept = (course_row.get('department') or '').strip()
            if course_dept:
                dept_row = conn.execute('SELECT id FROM departments WHERE LOWER(TRIM(name)) = LOWER(TRIM(?)) LIMIT 1', (course_dept,)).fetchone()
                if dept_row:
                    department_id = dept_row['id']

        # Final fallback: use current user's department when allowed
        if not department_id:
            department_id = current_department_id()

        # Enforce department scoping: non-super admins can only operate within their department
        if not is_super_admin():
            if department_id != current_department_id():
                conn.rollback()
                return None, 'غير مسموح لك بإجراء عمليات خارج قسمك.'

        submitted_start_time = (data.get('start_time') or '').strip() or None
        submitted_end_time = (data.get('end_time') or '').strip() or None

        # Determine final storage (NULL if default period time)
        period_lookup = _period_map(conn)
        final_start_time = submitted_start_time
        final_end_time = submitted_end_time
        p_details = period_lookup.get(period_code)
        if p_details and submitted_start_time == p_details['start_time'] and submitted_end_time == p_details['end_time']:
            final_start_time = None
            final_end_time = None

        # Allow caller to force save (bypass conflict detection) by passing `force_save` in the form data
        force_save_flag = str(data.get('force_save') or '').lower() in {'1', 'on', 'true', 'yes'}
        
        conflict_error = None
        if not force_save_flag:
            conflict_error = _validate_schedule_conflicts(
                conn, day=day, semester=semester, period_code=period_code, department_id=department_id,
                teacher_id=teacher_id, room_id=room_id, exclude_entry_id=entry_id,
                start_time=submitted_start_time, end_time=submitted_end_time
            )
            if conflict_error:
                conn.rollback()
                return None, conflict_error

            if entry_id:  # EDIT MODE
                old_entry = _entry_with_relations(conn, entry_id)
                if not old_entry:
                    conn.rollback()
                    return None, 'الحصة غير موجودة.'
                conn.execute(
                    'UPDATE timetable SET user_id=?, day=?, semester=?, section=?, course_id=?, teacher_id=?, room_id=?, start_time=?, end_time=?, department_id=? WHERE id=?',
                    (owner_user_id, day, semester, period_code, course_id, teacher_id, room_id, final_start_time, final_end_time, department_id, entry_id)
                )
                updated = _entry_with_relations(conn, entry_id)
                db.add_history(conn, 'EDIT', 'timetable', entry_id, *_history_actor(), message=_timetable_history_message('EDIT', dict(updated)), old_value=dict(old_entry), new_value=dict(updated))
                return updated, None
            else:  # CREATE MODE
                # Prevent DB-level UNIQUE constraint from raising by checking for
                # an existing timetable row with same user/day/semester/section.
                existing = conn.execute(
                    'SELECT id FROM timetable WHERE user_id=? AND day=? AND semester=? AND section=?',
                    (owner_user_id, day, semester, period_code)
                ).fetchone()

                if existing:
                    # Extract id from row object
                    existing_id = None
                    try:
                        existing_id = existing['id']
                    except Exception:
                        try:
                            existing_id = existing[0]
                        except Exception:
                            existing_id = None

                    # If caller requested a force save, replace the existing entry
                    if force_save_flag and existing_id:
                        conn.execute(
                            'UPDATE timetable SET user_id=?, day=?, semester=?, section=?, course_id=?, teacher_id=?, room_id=?, start_time=?, end_time=?, department_id=? WHERE id=?',
                            (owner_user_id, day, semester, period_code, course_id, teacher_id, room_id, final_start_time, final_end_time, department_id, existing_id)
                        )
                        updated = _entry_with_relations(conn, existing_id)
                        db.add_history(conn, 'EDIT', 'timetable', existing_id, *_history_actor(), message=_timetable_history_message('EDIT', dict(updated)), old_value=None, new_value=dict(updated))
                        return updated, None
                    else:
                        conn.rollback()
                        # Signal special redirect to caller with existing id when available
                        if existing_id:
                            return None, f'EXISTING:{existing_id}'
                        return None, 'يوجد صف بالفعل بنفس المستخدم، اليوم، والترتيب (القسم). استخدم التعديل أو فعّل Force Save لتجاوز ذلك.'

                cursor = conn.execute(
                    'INSERT INTO timetable (user_id, day, semester, section, course_id, teacher_id, room_id, start_time, end_time, department_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                    (owner_user_id, day, semester, period_code, course_id, teacher_id, room_id, final_start_time, final_end_time, department_id)
                )
                new_id = cursor.lastrowid
                added = _entry_with_relations(conn, new_id)
                db.add_history(conn, 'ADD', 'timetable', new_id, *_history_actor(), message=_timetable_history_message('ADD', dict(added)), new_value=dict(added))
                return added, None
        # If caller requested force_save we still need to perform the update/insert
        # operations but skip conflict detection above. Ensure we return a tuple
        # so callers can unpack the result.
        if force_save_flag:
            if entry_id:  # EDIT MODE (force replace)
                old_entry = _entry_with_relations(conn, entry_id)
                if not old_entry:
                    conn.rollback()
                    return None, 'الحصة غير موجودة.'
                conn.execute(
                    'UPDATE timetable SET user_id=?, day=?, semester=?, section=?, course_id=?, teacher_id=?, room_id=?, start_time=?, end_time=?, department_id=? WHERE id=?',
                    (owner_user_id, day, semester, period_code, course_id, teacher_id, room_id, final_start_time, final_end_time, department_id, entry_id)
                )
                updated = _entry_with_relations(conn, entry_id)
                db.add_history(conn, 'EDIT', 'timetable', entry_id, *_history_actor(), message=_timetable_history_message('EDIT', dict(updated)), old_value=dict(old_entry), new_value=dict(updated))
                return updated, None
            else:  # CREATE MODE (force insert or replace existing)
                existing = conn.execute(
                    'SELECT id FROM timetable WHERE user_id=? AND day=? AND semester=? AND section=?',
                    (owner_user_id, day, semester, period_code)
                ).fetchone()
                existing_id = None
                if existing:
                    try:
                        existing_id = existing['id']
                    except Exception:
                        try:
                            existing_id = existing[0]
                        except Exception:
                            existing_id = None

                if existing_id:
                    conn.execute(
                        'UPDATE timetable SET user_id=?, day=?, semester=?, section=?, course_id=?, teacher_id=?, room_id=?, start_time=?, end_time=?, department_id=? WHERE id=?',
                        (owner_user_id, day, semester, period_code, course_id, teacher_id, room_id, final_start_time, final_end_time, department_id, existing_id)
                    )
                    updated = _entry_with_relations(conn, existing_id)
                    db.add_history(conn, 'EDIT', 'timetable', existing_id, *_history_actor(), message=_timetable_history_message('EDIT', dict(updated)), old_value=None, new_value=dict(updated))
                    return updated, None

                cursor = conn.execute(
                    'INSERT INTO timetable (user_id, day, semester, section, course_id, teacher_id, room_id, start_time, end_time, department_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                    (owner_user_id, day, semester, period_code, course_id, teacher_id, room_id, final_start_time, final_end_time, department_id)
                )
                new_id = cursor.lastrowid
                added = _entry_with_relations(conn, new_id)
                db.add_history(conn, 'ADD', 'timetable', new_id, *_history_actor(), message=_timetable_history_message('ADD', dict(added)), new_value=dict(added))
                return added, None
    except Exception as e:
        conn.rollback()
        raise e


def _ensure_selection(records, selected_id, conn, table_name):
    if selected_id and not any(item['id'] == selected_id for item in records):
        row = conn.execute(f'SELECT * FROM {table_name} WHERE id = ?', (selected_id,)).fetchone()
        if row:
            records.insert(0, dict(row))
    return records


def _room_availability_rows(conn, day, semester, period_code, exclude_entry_id=None, start_time=None, end_time=None, include_conflict_details=True):
    rooms = _fetch_rooms(conn)
    if not rooms:
        return []

    period_lookup = _period_map(conn)
    target_start, target_end = _resolve_time_range(period_lookup, period_code, start_time, end_time)
    
    # Fetch all entries for this day/semester to check overlaps
    sql = """
        SELECT t.room_id, c.name as course_name, d.name as dept_name, t.section, t.start_time, t.end_time, t.id
        FROM timetable t
        JOIN courses c ON c.id = t.course_id
        LEFT JOIN departments d ON d.id = t.department_id
        WHERE t.day = ? AND t.semester = ?
    """
    entries = conn.execute(sql, (day, semester)).fetchall()
    
    conflicts = {}
    for entry in entries:
        if exclude_entry_id is not None and entry['id'] == exclude_entry_id:
            continue
        
        e_start, e_end = _resolve_time_range(period_lookup, entry['section'], entry['start_time'], entry['end_time'])
        if _time_ranges_overlap(target_start, target_end, e_start, e_end):
            conflicts[entry['room_id']] = f"محجوزة بواسطة {entry['dept_name']} ({entry['course_name']})"

    if not target_start or not target_end:
        return [{**dict(room), 'is_available': True} for room in rooms]

    room_rows = []
    for room in rooms:
        room_dict = dict(room)
        conflict_msg = conflicts.get(room_dict['id'])
        room_dict['is_available'] = conflict_msg is None
        room_dict['conflict_msg'] = conflict_msg
        room_rows.append(room_dict)
    return room_rows


def _available_teachers_detailed(conn, day, semester, period_code, owner_user_id, exclude_entry_id=None, start_time=None, end_time=None, department_name=None):
    teachers = _fetch_teachers(
        conn,
        owner_user_id=owner_user_id,
        include_all=is_super_admin(),
        department_name=department_name if is_super_admin() else _scoped_department_name(conn),
    )
    if not teachers and not is_super_admin():
        teachers = _fetch_teachers(conn, owner_user_id=owner_user_id, include_all=False, department_name=None)
    
    period_lookup = _period_map(conn)
    target_start, target_end = _resolve_time_range(period_lookup, period_code, start_time, end_time)
    
    sql = """
        SELECT t.teacher_id, c.name as course_name, t.semester as sem, t.section, t.start_time, t.end_time, t.id, d.name as dept_name
        FROM timetable t
        JOIN courses c ON c.id = t.course_id
        LEFT JOIN departments d ON d.id = t.department_id
        WHERE t.day = ? AND t.semester = ?
    """
    entries = conn.execute(sql, (day, semester)).fetchall()
    
    conflicts = {}
    for entry in entries:
        if exclude_entry_id is not None and entry['id'] == exclude_entry_id:
            continue
        
        e_start, e_end = _resolve_time_range(period_lookup, entry['section'], entry['start_time'], entry['end_time'])
        if _time_ranges_overlap(target_start, target_end, e_start, e_end):
            conflicts[entry['teacher_id']] = f"مشغول مع {entry['course_name']} (فصل {entry['sem']}) — {(entry['dept_name'] or '')}"

    results = []
    for t in teachers:
        t_dict = dict(t)
        conflict_msg = conflicts.get(t_dict['id'])
        t_dict['is_available'] = conflict_msg is None
        t_dict['conflict_msg'] = conflict_msg
        results.append(t_dict)
    return results


def _available_teachers(conn, day, semester, period_code, owner_user_id, exclude_entry_id=None, start_time=None, end_time=None, department_name=None):
    """Simple wrapper that returns available teachers using the detailed implementation.

    Kept for backward compatibility with places that call `_available_teachers`.
    """
    return _available_teachers_detailed(
        conn,
        day,
        semester,
        period_code,
        owner_user_id,
        exclude_entry_id=exclude_entry_id,
        start_time=start_time,
        end_time=end_time,
        department_name=department_name,
    )


@timetable_bp.route('/available-rooms')
@timetable_bp.route('/api/available-rooms')
@login_required
@courses_timetable_admin_required
def available_rooms():
    conn = db.get_db()
    day = (request.args.get('day') or '').strip()
    semester = _safe_int(request.args.get('semester'))
    period_code = (request.args.get('section') or request.args.get('period_code') or '').strip()
    exclude_id = _safe_int(request.args.get('exclude_id'))
    department_id = _safe_int(request.args.get('department_id'))
    if not day or semester is None or not period_code:
        return jsonify({'rooms': [], 'error': 'day, semester, period_code are required'}), 400
    return jsonify({'rooms': _room_availability_rows(conn, day, semester, period_code, exclude_entry_id=exclude_id, start_time=request.args.get('start_time'), end_time=request.args.get('end_time'))})


@timetable_bp.route('/available-teachers')
@timetable_bp.route('/api/available-teachers')
@login_required
@courses_timetable_admin_required
def available_teachers():
    conn = db.get_db()
    day = (request.args.get('day') or '').strip()
    semester = _safe_int(request.args.get('semester'))
    period_code = (request.args.get('section') or request.args.get('period_code') or '').strip()
    exclude_id = _safe_int(request.args.get('exclude_id'))
    if not day or semester is None or not period_code:
        return jsonify({'teachers': [], 'error': 'day, semester, period_code are required'}), 400
    dept_name = _selected_department_name(conn)
    department_id = _safe_int(request.args.get('department_id')) or _selected_department_id()
    teachers = _available_teachers(conn, day, semester, period_code, current_user_id(), exclude_entry_id=exclude_id, start_time=request.args.get('start_time'), end_time=request.args.get('end_time'), department_name=dept_name)
    return jsonify({'teachers': teachers})


@timetable_bp.route('/form', defaults={'id': None}, methods=['GET', 'POST'])
@timetable_bp.route('/<int:id>/form', methods=['GET', 'POST'])
@login_required
@courses_timetable_admin_required
def timetable_form(id):
    conn = db.get_db()
    entry = None
    if id:
        entry = _entry_with_relations(conn, id)
        if not entry:
            flash('الحصة غير موجودة.', 'danger')
            return redirect(url_for('timetable.list_timetable'))
        if not _can_access_entry(conn, entry):
            flash('غير مسموح لك بتعديل هذه الحصة.', 'danger')
            return redirect(url_for('timetable.list_timetable'))

    is_modal = request.values.get('modal') in {'1', 'true', 'True'}
    next_url = (request.values.get('next') or '').strip()
    
    # Mode detection for processing
    if request.method == 'POST':
        result, error = _process_timetable_save(conn, id, request.form)
        if error:
            # Special signal: EXISTING:<id> -> redirect user to edit that existing entry
            if isinstance(error, str) and error.startswith('EXISTING:'):
                try:
                    existing_id = int(error.split(':', 1)[1])
                except Exception:
                    flash('يوجد صف متعارض بالفعل.', 'warning')
                    return redirect(request.url)
                flash('يوجد صف موجود لنفس المستخدم؛ افتح التعديل لضبطه أو فعّل Force Save.', 'warning')
                return redirect(url_for('timetable.timetable_form', id=existing_id))

            flash(error, 'danger')
            return redirect(request.url)

        conn.commit()
        msg = 'تم تحديث الحصة بنجاح' if id else 'تم حفظ الحصة بنجاح'
        flash(msg, 'success')
        if is_modal:
            return render_template('timetable/modal_success.html', message=msg)
        return redirect(next_url or url_for('timetable.list_timetable'))

    # GET REQUEST - Render Form
    enabled_periods = _get_periods(conn, enabled_only=True)

    # Pre-fill logic
    # Determine selected department early so we can pick a sensible default semester
    dept_id = _selected_department_id() or (entry.get('department_id') if entry else None)

    # Default semester behavior:
    # - If editing an entry, preserve its semester.
    # - Otherwise, if the selected department has more than one semester, default to semester 2.
    default_sem_fallback = entry['semester'] if entry else 1
    if not entry and dept_id and not is_super_admin():
        dept_info = _department_id_map(conn).get(dept_id)
        dept_semesters = 0
        if dept_info and dept_info.get('semesters') and str(dept_info.get('semesters')).isdigit():
            try:
                dept_semesters = int(dept_info.get('semesters'))
            except Exception:
                dept_semesters = 0
        # Prefer semester 2 by default when the department actually has multiple semesters.
        if dept_semesters > 1:
            default_sem_fallback = 2

    d_sem = _safe_int(request.values.get('semester'), default_sem_fallback)
    required_year = (d_sem + 1) // 2
    d_day = request.args.get('day', entry['day'] if entry else DAY_NAMES[0])
    d_period = request.args.get('section', entry['section'] if entry else (enabled_periods[0]['code'] if enabled_periods else 'A'))

    f_start = entry['start_time'] if entry else None
    f_end = entry['end_time'] if entry else None

    # NEW: Standardized Department Filtering
    # capture selected department name as well for form rendering
    if is_super_admin():
        dept_name = request.args.get('department_name') or _selected_department_name(conn)
    else:
        dept_name = _scoped_department_name(conn)
    
    dept_filter = (dept_name or '').strip()

    all_dept_courses = _fetch_courses(conn, include_all=is_super_admin(), owner_user_id=current_user_id(), department_name=dept_filter)
    courses = [c for c in all_dept_courses if c['year'] == required_year]
    
    # Debug Info
    print(f"[DEBUG] Sem: {d_sem} | Yr: {required_year} | Dept: {dept_filter} | Courses: {len(courses)}/{len(all_dept_courses)}")

    initial_rooms = _room_availability_rows(conn, d_day, d_sem, d_period, exclude_entry_id=id, start_time=f_start, end_time=f_end)
    initial_teachers = _available_teachers(conn, d_day, d_sem, d_period, current_user_id(), exclude_entry_id=id, start_time=f_start, end_time=f_end, department_name=dept_name)

    return render_template(
        'timetable/form.html',
        entry=entry,
        days=DAY_NAMES,
        enabled_periods=enabled_periods,
        courses=courses,
        teachers=initial_teachers,
        initial_rooms=initial_rooms,
        default_day=d_day,
        default_semester=d_sem,
        default_period_code=d_period,
        form_start_time=f_start,
        form_end_time=f_end,
        is_modal=is_modal,
        next_url=next_url,
        selected_department_id=dept_id,
        selected_department_name=dept_name
        , hide_nav=is_modal
    )


@timetable_bp.route('/api/report-error', methods=['POST'])
@login_required
def report_error():
    payload = request.get_json(silent=True) or {}
    msg = payload.get('message', 'No message')
    err_url = payload.get('url', 'No URL')
    
    conn = db.get_db()
    db.add_history(
        conn,
        'SYSTEM_ERROR',
        'modal_iframe',
        entity_id=None,
        actor_user_id=current_user_id(),
        actor_username=session.get('username') or session.get('label'),
        message=f"User reported a 500 error in modal. URL: {err_url} | Info: {msg}",
        new_value=payload
    )
    conn.commit()
    return jsonify({'ok': True})


@timetable_bp.route('/period-settings', methods=['POST'])
@login_required
@super_admin_required
def update_period_settings():
    conn = db.get_db()
    is_modal = request.form.get('modal') in {'1', 'true', 'True'}
    periods, form_error = _periods_from_form(request.form)
    if form_error:
        flash(form_error, 'danger')
        return redirect(url_for('timetable.list_timetable'))
    validation_error = _validate_period_definitions(periods)
    if validation_error:
        flash(validation_error, 'danger')
        return redirect(url_for('timetable.list_timetable'))
    for period in periods:
        conn.execute('UPDATE period_settings SET label = ?, start_time = ?, end_time = ?, is_enabled = ?, sort_order = ? WHERE code = ?', (period['label'], period['start_time'], period['end_time'], period['is_enabled'], period['sort_order'], period['code']))
    conn.commit()
    flash('Period settings saved successfully.', 'success')
    if is_modal:
        return render_template('timetable/modal_success.html', message='تم تحديث إعدادات الفترات')
    return redirect(url_for('timetable.list_timetable'))


@timetable_bp.route('/')
@login_required
@courses_timetable_admin_required
@admin_department_required
def list_timetable():
    conn = db.get_db()
    all_departments = _fetch_departments(conn)
    dept_id_map = _department_id_map(conn)
    department_semester_map = {d['name']: d['semesters'] for d in all_departments}

    if is_super_admin():
        # Allow super admins to view all departments by leaving selection empty.
        # Only use an explicit department_id when provided in the request.
        selected_department_id = _selected_department_id()
    else:
        selected_department_id = current_department_id()

    dept_info = dept_id_map.get(selected_department_id)
    selected_department_name = dept_info['name'] if dept_info else ''

    # Compute allowed semesters: respect department configuration when a specific
    # department is selected; otherwise super_admin may see the full range.
    if selected_department_name:
        allowed_semesters = _build_allowed_semesters(selected_department_name, department_semester_map)
    else:
        # No specific department selected: super_admins can view full semester range
        if is_super_admin():
            allowed_semesters = list(range(1, 9))
        else:
            allowed_semesters = _build_allowed_semesters(selected_department_name, department_semester_map)
    
    # Support a single-select `semester` plus an optional `show_all` toggle.
    if 'semester' in request.args or 'show_all' in request.args:
        # When user explicitly interacts, update session.
        if 'semester' in request.args:
            # single value or multiple values both handled by getlist
            selected_semesters_raw = request.args.getlist('semester')
        else:
            selected_semesters_raw = []
        # If show_all is present and truthy, select all allowed semesters
        show_all_flag = str(request.args.get('show_all') or '').lower() in {'1', 'on', 'true', 'yes'}
        if show_all_flag:
            selected_semesters = list(allowed_semesters)
            session['selected_semesters'] = [str(s) for s in selected_semesters]
        else:
            selected_semesters = [int(value) for value in selected_semesters_raw if str(value).isdigit() and int(value) in allowed_semesters]
            session['selected_semesters'] = [str(s) for s in selected_semesters]
    else:
        selected_semesters_raw = session.get('selected_semesters', [])
        selected_semesters = [int(value) for value in selected_semesters_raw if str(value).isdigit() and int(value) in allowed_semesters]
    # If the user didn't pass semester explicitly and the session contained a value,
    # allow override for super_admin when a specific department is selected: prefer semester 2.
            # Compute allowed semesters using configured department semesters (no automatic
            # expansion for super_admin). UI-level controls already differ per role.
            allowed_semesters = _build_allowed_semesters(selected_department_name, department_semester_map)

        # If a specific department is selected and it supports multiple semesters,
        # prefer semester 2. Do not force semester 2 for single-semester departments.
        if selected_department_id and dept_semesters > 1 and 2 in allowed_semesters:
            selected_semesters = [2]
            session['selected_semesters'] = ['2']
        else:
            selected_semesters = [allowed_semesters[0]] if allowed_semesters else [1]
    else:
        # There may be a previously stored session selection. If a specific department is
        # selected and that department actually supports more than one semester, default
        # to semester 2 when the user didn't explicitly pass a semester in the request.
        dept_info = dept_id_map.get(selected_department_id)
        dept_semesters = 0
        if dept_info and dept_info.get('semesters') and str(dept_info.get('semesters')).isdigit():
            try:
                dept_semesters = int(dept_info.get('semesters'))
            except Exception:
                dept_semesters = 0

        if selected_department_id and dept_semesters > 1 and 'semester' not in request.args and 2 in allowed_semesters:
            selected_semesters = [2]
            session['selected_semesters'] = ['2']
        else:
            session['selected_semesters'] = [str(s) for s in selected_semesters]

    search = (request.args.get('q') or '').strip()
    enabled_periods = _get_periods(conn, enabled_only=True)
    sql = """
        SELECT t.*, c.name AS course_name, c.department AS course_department,
               te.name AS teacher_name,
               r.name AS room_name, r.name_ar AS room_name_ar,
               u.username AS owner_username, u.label AS owner_label
        FROM timetable t
        JOIN courses c ON c.id = t.course_id
        JOIN teachers te ON te.id = t.teacher_id
        JOIN rooms r ON r.id = t.room_id
        LEFT JOIN users u ON u.id = t.user_id
        WHERE t.semester IN ({})
    """.format(','.join('?' for _ in selected_semesters))
    params = list(selected_semesters)

    if selected_department_id:
        # Some timetable rows store department by id, others only by course.department name.
        # Match either the stored `department_id` or the course's `department` name (case-insensitive, trimmed).
        sql += ' AND (t.department_id = ? OR LOWER(TRIM(c.department)) = LOWER(TRIM(?)))'
        params.append(selected_department_id)
        params.append(selected_department_name or '')

    if search:
        sql += ' AND (te.name LIKE ? OR r.name LIKE ? OR r.name_ar LIKE ? OR c.name LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'])

    sql += ' ORDER BY t.day, t.semester, t.section, t.created_at'
    rows = conn.execute(sql, params).fetchall()

    # FALLBACK for export too: if department selected and no rows, try matching by
    # course.department name or department admin owners (legacy data).
    if selected_department_id and not rows:
        try:
            fallback_sql = """
                SELECT t.*, c.name AS course_name, c.department AS course_department, te.name AS teacher_name, r.name AS room_name, r.name_ar AS room_name_ar
                FROM timetable t
                JOIN courses c ON c.id = t.course_id
                JOIN teachers te ON te.id = t.teacher_id
                JOIN rooms r ON r.id = t.room_id
                WHERE t.semester IN ({}) AND (
                    LOWER(TRIM(c.department)) = LOWER(TRIM(?))
                    OR t.user_id IN (SELECT id FROM users WHERE role = 'admin' AND department_id = ?)
                )
            """.format(','.join('?' for _ in selected_semesters))
            fb_params = list(selected_semesters)
            fb_params.append(selected_department_name or '')
            fb_params.append(selected_department_id)
            rows = conn.execute(fallback_sql, fb_params).fetchall()
        except Exception:
            pass

    # FALLBACK: If a specific department was requested but no rows returned,
    # try to fetch rows by matching course.department name or entries owned by
    # department admin users. This addresses legacy rows where t.department_id
    # is NULL but c.department is populated.
    if selected_department_id and not rows:
        try:
            fallback_sql = """
                SELECT t.*, c.name AS course_name, c.department AS course_department,
                       te.name AS teacher_name, r.name AS room_name, r.name_ar AS room_name_ar
                FROM timetable t
                JOIN courses c ON c.id = t.course_id
                JOIN teachers te ON te.id = t.teacher_id
                JOIN rooms r ON r.id = t.room_id
                WHERE t.semester IN ({}) AND (
                    LOWER(TRIM(c.department)) = LOWER(TRIM(?))
                    OR t.user_id IN (SELECT id FROM users WHERE role = 'admin' AND department_id = ?)
                )
            """.format(','.join('?' for _ in selected_semesters))
            fb_params = list(selected_semesters)
            fb_params.append(selected_department_name or '')
            fb_params.append(selected_department_id)
            rows = conn.execute(fallback_sql, fb_params).fetchall()
        except Exception:
            pass

    timetable = {}
    for row in rows:
        day_map = timetable.setdefault(row['day'], {})
        # Normalize semester key to int for template compatibility
        try:
            sem_key = int(row['semester'])
        except Exception:
            sem_key = row['semester']
        sem_map = day_map.setdefault(sem_key, {})
        # ALWAYS use list for slot structure (Clean pattern to prevent duplication)
        slot = sem_map.setdefault(row['section'], [])
        slot.append(dict(row))

    current_owner_id = current_user_id()
    department_filter = selected_department_name if is_super_admin() else _scoped_department_name(conn)
    courses = _fetch_courses(conn, include_all=is_super_admin(), owner_user_id=current_owner_id, department_name=department_filter)
    teachers = _fetch_teachers(conn, owner_user_id=current_owner_id, include_all=is_super_admin(), department_name=department_filter)
    rooms = _fetch_rooms(conn)
    period_settings = _get_periods(conn, enabled_only=False)

    return render_template(
        'timetable/list.html',
        timetable=timetable,
        days=DAY_NAMES,
        day_labels=DAY_LABELS,
        enabled_periods=enabled_periods,
        period_settings=period_settings,
        selected_department_id=selected_department_id,
        selected_semesters=selected_semesters,
        allowed_semesters=allowed_semesters,
        departments=all_departments,
        all_departments=all_departments,
        search=search,
        selected_department_name=selected_department_name,
        users=[],
        courses=courses,
        teachers=teachers,
        rooms=rooms,
        can_delete_critical=_can_delete_critical(),
    )


@timetable_bp.route('/export')
@login_required
@courses_timetable_admin_required
@admin_department_required
def export_timetable():
    """Generates an Excel report of the timetable, respecting active filters and search query."""
    conn = db.get_db()
    
    # Identify active filters (same logic as list_timetable)
    all_departments = _fetch_departments(conn)
    dept_id_map = _department_id_map(conn)
    department_semester_map = {d['name']: d['semesters'] for d in all_departments}

    if is_super_admin():
        # Do not default to the first department for super admins; None means "all departments".
        selected_department_id = _selected_department_id()
    else:
        selected_department_id = current_department_id()

    dept_info = dept_id_map.get(selected_department_id)
    selected_department_name = dept_info['name'] if dept_info else ''

    allowed_semesters = _build_allowed_semesters(selected_department_name, department_semester_map)
    selected_semesters_raw = request.args.getlist('semester') or session.get('selected_semesters', [])
    selected_semesters = [int(v) for v in selected_semesters_raw if str(v).isdigit() and int(v) in allowed_semesters]
    if not selected_semesters:
        selected_semesters = [allowed_semesters[0]] if allowed_semesters else [1]

    search = (request.args.get('q') or '').strip()

    # Build query with identical search/filter logic
    sql = """
        SELECT t.*, c.name AS course_name, c.department AS course_department, te.name AS teacher_name, 
               r.name AS room_name, r.name_ar AS room_name_ar
        FROM timetable t
        JOIN courses c ON c.id = t.course_id
        JOIN teachers te ON te.id = t.teacher_id
        JOIN rooms r ON r.id = t.room_id
        WHERE t.semester IN ({})
    """.format(','.join('?' for _ in selected_semesters))
    params = list(selected_semesters)

    if selected_department_id:
        sql += ' AND (t.department_id = ? OR LOWER(TRIM(c.department)) = LOWER(TRIM(?)))'
        params.append(selected_department_id)
        params.append(selected_department_name or '')
    if search:
        sql += ' AND (te.name LIKE ? OR r.name LIKE ? OR r.name_ar LIKE ? OR c.name LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'])

    sql += ' ORDER BY t.day, t.semester, t.section, t.created_at'
    rows = conn.execute(sql, params).fetchall()

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير الجدول الدراسي"
    ws.sheet_view.rightToLeft = True

    # Header configuration
    headers = ['اليوم', 'الفصل', 'الفترة', 'المقرر', 'المدرس', 'القاعة', 'الوقت', 'اقتراح الحل']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    conflict_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # Light red for conflicts
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center')
    overload_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    overload_font = Font(color='9C0006', bold=True)
    WORKLOAD_THRESHOLD = 18.0

    # Pre-calculate slot occupancy to identify conflicts
    slot_counts = {}
    for r in rows:
        key = (r['day'], r['semester'], r['section'])
        slot_counts[key] = slot_counts.get(key, 0) + 1

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Populate Rows
    period_map = _period_map(conn)
    enabled_periods = _get_periods(conn, enabled_only=True)

    for idx, r in enumerate(rows, 2):
        p_info = period_map.get(r['section'])
        time_val = f"{r['start_time'] or p_info['start_time']} ← {r['end_time'] or p_info['end_time']}" if p_info else ""
        
        # Comprehensive conflict detection
        semester_overlap = slot_counts.get((r['day'], r['semester'], r['section']), 0) > 1
        conflict_entry = check_timetable_conflict(conn, r['course_department'], r['day'], r['semester'], r['section'], exclude_id=r['id'])
        db_conflict_msg = f"تعارض مع: {conflict_entry['course_name']}" if conflict_entry else None
        is_conflict = semester_overlap or bool(db_conflict_msg)

        ws.cell(row=idx, column=1, value=r['day'])
        ws.cell(row=idx, column=2, value=r['semester'])
        ws.cell(row=idx, column=3, value=p_info['label'] if p_info else r['section'])
        ws.cell(row=idx, column=4, value=r['course_name'])
        ws.cell(row=idx, column=5, value=r['teacher_name'])
        ws.cell(row=idx, column=6, value=r['room_name_ar'] or r['room_name'])
        ws.cell(row=idx, column=7, value=time_val)
        
        # Conflict Resolution Suggestion logic
        resolution_msg = ""
        if is_conflict:
            suggestions = []
            if db_conflict_msg:
                suggestions.append(f"تنبيه: {db_conflict_msg}")
            elif semester_overlap:
                suggestions.append("تنبيه: يوجد تداخل في حصص الفصل الدراسي")

            # 1. Suggest available rooms in the same period
            avail_rooms = _room_availability_rows(conn, r['day'], r['semester'], r['section'], exclude_entry_id=r['id'])
            free_rooms = [rm['name_ar'] or rm['name'] for rm in avail_rooms if rm['is_available']]
            if free_rooms:
                suggestions.append(f"قاعات متاحة: {', '.join(free_rooms[:2])}")

            # 2. Suggest alternative periods on the same day for this teacher and room
            free_slots = [p['label'] for p in enabled_periods if p['code'] != r['section'] and not _validate_schedule_conflicts(conn, day=r['day'], semester=r['semester'], period_code=p['code'], teacher_id=r['teacher_id'], room_id=r['room_id'])]
            if free_slots:
                suggestions.append(f"أوقات بديلة: {', '.join(free_slots)}")

            resolution_msg = " | ".join(suggestions) if suggestions else "لا يوجد حل مباشر متاح"

        ws.cell(row=idx, column=8, value=resolution_msg)

        for col in range(1, 9):
            cell = ws.cell(row=idx, column=col)
            cell.alignment = center_align
            if is_conflict:
                cell.fill = conflict_fill

    # Auto-adjust column width
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 5

    # Create Summary Sheet
    ws_sum = wb.create_sheet(title="إحصائيات إجمالية")
    ws_sum.sheet_view.rightToLeft = True

    teacher_stats = {}
    room_stats = {}

    for r in rows:
        p_info = period_map.get(r['section'])
        st_str = r['start_time'] or (p_info['start_time'] if p_info else None)
        et_str = r['end_time'] or (p_info['end_time'] if p_info else None)
        
        duration = 0
        if st_str and et_str:
            try:
                st_obj = _parse_time(st_str)
                et_obj = _parse_time(et_str)
                duration = (_minutes(et_obj) - _minutes(st_obj)) / 60.0
            except (ValueError, TypeError):
                duration = 0

        t_name = r['teacher_name']
        teacher_stats[t_name] = teacher_stats.get(t_name, 0) + duration

        r_name = r['room_name_ar'] or r['room_name']
        room_stats[r_name] = room_stats.get(r_name, 0) + duration

    # Section: Teachers
    ws_sum.merge_cells('A1:B1')
    ws_sum['A1'] = "إحصائيات المدرسين (بالساعات)"
    ws_sum['A1'].font = Font(bold=True, size=12)
    ws_sum['A1'].alignment = Alignment(horizontal='center')

    ws_sum.cell(row=2, column=1, value="اسم المدرس").font = header_font
    ws_sum.cell(row=2, column=1).fill = header_fill
    ws_sum.cell(row=2, column=2, value="إجمالي الساعات").font = header_font
    ws_sum.cell(row=2, column=2).fill = header_fill

    curr = 3
    for name, hrs in sorted(teacher_stats.items()):
        c1 = ws_sum.cell(row=curr, column=1, value=name)
        c2 = ws_sum.cell(row=curr, column=2, value=hrs)
        c1.alignment = center_align
        c2.alignment = center_align
        
        if hrs > WORKLOAD_THRESHOLD:
            c1.fill = overload_fill
            c1.font = overload_font
            c2.fill = overload_fill
            c2.font = overload_font
        curr += 1

    # Section: Rooms (with a small gap)
    curr += 2
    ws_sum.merge_cells(f'A{curr}:B{curr}')
    ws_sum.cell(row=curr, column=1, value="إحصائيات القاعات (بالساعات)").font = Font(bold=True, size=12)
    ws_sum.cell(row=curr, column=1).alignment = Alignment(horizontal='center')
    curr += 1

    ws_sum.cell(row=curr, column=1, value="اسم القاعة").font = header_font
    ws_sum.cell(row=curr, column=1).fill = header_fill
    ws_sum.cell(row=curr, column=2, value="إجمالي الساعات").font = header_font
    ws_sum.cell(row=curr, column=2).fill = header_fill
    curr += 1

    for name, hrs in sorted(room_stats.items()):
        ws_sum.cell(row=curr, column=1, value=name).alignment = center_align
        ws_sum.cell(row=curr, column=2, value=hrs).alignment = center_align
        curr += 1

    ws_sum.column_dimensions['A'].width = 35
    ws_sum.column_dimensions['B'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name='timetable_report.xlsx', 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@timetable_bp.route('/create', methods=['GET', 'POST'])
@login_required
@courses_timetable_admin_required
def create_timetable():
    conn = db.get_db()
    enabled_periods = _get_periods(conn, enabled_only=True)
    default_day = request.args.get('day', DAY_NAMES[0])
    # Determine selected department and its semester count to choose a sensible default semester
    selected_department_id = _selected_department_id()
    dept_info = _department_id_map(conn).get(selected_department_id)
    dept_semesters = 0
    if dept_info and dept_info.get('semesters') and str(dept_info.get('semesters')).isdigit():
        try:
            dept_semesters = int(dept_info.get('semesters'))
        except Exception:
            dept_semesters = 0
    default_semester = _safe_int(request.args.get('semester'), 2 if dept_semesters > 1 else 1)
    default_period_code = request.args.get('section', enabled_periods[0]['code'] if enabled_periods else 'A')
    is_modal = request.values.get('modal') in {'1', 'true', 'True'}

    # Fetch default start_time and end_time for the selected period
    period_lookup = _period_map(conn)
    default_period_details = period_lookup.get(default_period_code)
    default_period_start_time = default_period_details['start_time'] if default_period_details else None
    default_period_end_time = default_period_details['end_time'] if default_period_details else None

    next_url = (request.values.get('next') or '').strip()
    selected_department_id = selected_department_id or _selected_department_id()
    selected_department_name = _selected_department_name(conn)

    if request.method == 'POST':
        result, error = _process_timetable_save(conn, None, request.form)
        if error:
            flash(error, 'danger')
            return redirect(next_url or url_for('timetable.create_timetable', department_id=selected_department_id))

        conn.commit()
        flash('Timetable entry created successfully.', 'success')
        if is_modal:
            return render_template('timetable/modal_success.html', message='تم حفظ الحصة بنجاح')
        return redirect(next_url or url_for('timetable.list_timetable'))

    initial_rooms = _room_availability_rows(conn, default_day, default_semester, default_period_code)
    initial_teachers = _available_teachers(conn, default_day, default_semester, default_period_code, current_user_id(), department_name=selected_department_name)
    department_filter = selected_department_name if is_super_admin() else _scoped_department_name(conn)

    return render_template(
        'timetable/form.html',
        days=DAY_NAMES,
        enabled_periods=enabled_periods,
        courses=_fetch_courses(conn, include_all=is_super_admin(), owner_user_id=current_user_id(), department_name=department_filter),
        teachers=_fetch_teachers(conn, owner_user_id=current_user_id(), include_all=is_super_admin(), department_name=department_filter),
        initial_rooms=initial_rooms,
        initial_teachers=initial_teachers,
        default_day=default_day,
        default_semester=default_semester,
        default_period_code=default_period_code,
        is_modal=is_modal,
        next_url=next_url,
        default_period_start_time=default_period_start_time, # Pass to template
        default_period_end_time=default_period_end_time,     # Pass to template
        selected_department_id=selected_department_id,
        selected_department_name=selected_department_name,
    )


@timetable_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@courses_timetable_admin_required
def edit_timetable(id):
    conn = db.get_db()
    entry = _entry_with_relations(conn, id)
    if not _can_access_entry(conn, entry):
        flash('You are not allowed to edit this entry.', 'danger')
        return redirect(url_for('timetable.list_timetable'))

    is_modal = request.values.get('modal') in {'1', 'true', 'True'}

    # Determine the times to pre-fill the form
    form_start_time = entry['start_time']
    form_end_time = entry['end_time']
    
    # If the entry itself doesn't have custom times, use the period's default times
    if not form_start_time or not form_end_time:
        period_lookup = _period_map(conn)
        period_details = period_lookup.get(entry['section'])
        if period_details:
            form_start_time = period_details['start_time']
            form_end_time = period_details['end_time']
    next_url = (request.values.get('next') or '').strip()

    # determine selected department context early (used by POST redirects and availability lookups)
    selected_department_id = _selected_department_id() or entry.get('department_id')
    selected_department_name = _selected_department_name(conn, fallback=entry.get('course_department'))

    if request.method == 'POST':
        result, error = _process_timetable_save(conn, id, request.form)
        if error:
            flash(error, 'danger')
            return redirect(next_url or url_for('timetable.edit_timetable', id=id, department_id=selected_department_id))

        conn.commit()
        flash('Timetable entry updated successfully.', 'success')
        if is_modal:
            return render_template('timetable/modal_success.html', message='تم تحديث الحصة بنجاح')
        return redirect(next_url or url_for('timetable.list_timetable'))

    enabled_periods = _get_periods(conn, enabled_only=True)
    available_rooms = _room_availability_rows(conn, entry['day'], entry['semester'], entry['section'], exclude_entry_id=id)
    available_teachers = _available_teachers(conn, entry['day'], entry['semester'], entry['section'], current_user_id(), exclude_entry_id=id, department_name=selected_department_name)
    available_teachers = _ensure_selection(available_teachers, entry['teacher_id'], conn, 'teachers')

    department_filter = selected_department_name if is_super_admin() else _scoped_department_name(conn)

    return render_template(
        'timetable/edit.html',
        entry=entry,
        days=DAY_NAMES,
        enabled_periods=enabled_periods,
        courses=_fetch_courses(conn, include_all=is_super_admin(), owner_user_id=current_user_id(), department_name=department_filter),
        teachers=_fetch_teachers(conn, owner_user_id=current_user_id(), include_all=is_super_admin(), department_name=department_filter),
        available_rooms=available_rooms,
        available_teachers=available_teachers,
        can_delete_critical=_can_delete_critical(),
        next_url=next_url,
        form_start_time=form_start_time, # Pass to template
        form_end_time=form_end_time,     # Pass to template
        selected_department_id=selected_department_id,
        selected_department_name=selected_department_name,
        is_modal=is_modal
    )


@timetable_bp.route('/api/quick-add', methods=['POST'])
@login_required
@courses_timetable_admin_required
def quick_add_from_cell():
    payload = request.get_json(silent=True) or {}
    conn = db.get_db()
    result, error = _process_timetable_save(conn, None, payload)
    if error:
        status = 409 if any(kw in error for kw in ['محجوزة', 'منشغل', 'لديك بالفعل']) else 400
        return jsonify({'ok': False, 'message': error}), status

    conn.commit()
    return jsonify({'ok': True, 'entry': dict(result), 'message': 'Timetable entry created successfully.'}), 201


@timetable_bp.route('/api/edit-entry', methods=['POST'])
@login_required
@courses_timetable_admin_required
def edit_entry():
    payload = request.get_json(silent=True) or {}
    conn = db.get_db()
    entry_id = _safe_int(payload.get('lecture_id'))
    entry = conn.execute('SELECT * FROM timetable WHERE id = ?', (entry_id,)).fetchone()
    if not _can_access_entry(conn, entry):
        return jsonify({'ok': False, 'message': 'Not allowed.'}), 403

    result, error = _process_timetable_save(conn, entry_id, payload)
    if error:
        status = 409 if any(kw in error for kw in ['محجوزة', 'منشغل', 'لديك بالفعل']) else 400
        return jsonify({'ok': False, 'message': error}), status

    conn.commit()
    return jsonify({'ok': True, 'entry': dict(result), 'message': 'Timetable entry updated successfully.'})


@timetable_bp.route('/api/delete-entry', methods=['POST'])
@login_required
@courses_timetable_admin_required
def delete_entry():
    payload = request.get_json(silent=True) or {}
    entry_id = _safe_int(payload.get('lecture_id'))
    conn = db.get_db()
    # fetch full entry with relations for reliable department info
    entry = _entry_with_relations(conn, entry_id)
    if not entry:
        return jsonify({'ok': False, 'message': 'الحصة غير موجودة.'}), 404

    # Permission: allow super_admin, or allow admin to delete entries within their department
    entry_dept_id = entry.get('department_id')
    if not entry_dept_id:
        # try to resolve from course department name
        course_dept = (entry.get('course_department') or '').strip()
        if course_dept:
            dept_row = conn.execute('SELECT id FROM departments WHERE LOWER(TRIM(name)) = LOWER(TRIM(?)) LIMIT 1', (course_dept,)).fetchone()
            if dept_row:
                entry_dept_id = dept_row['id']

    if not (is_super_admin() or (current_role() == 'admin' and entry_dept_id == current_department_id())):
        return jsonify({'ok': False, 'message': 'فقط المشرف العام أو أدمن القسم يمكنه حذف الحصص داخل قسمه.'}), 403

    if not _can_access_entry(conn, entry):
        return jsonify({'ok': False, 'message': 'Not allowed.'}), 403

    old_entry = entry
    db.add_history(
        conn,
        'DELETE',
        'timetable',
        entry_id,
        *_history_actor(),
        message=_timetable_history_message('DELETE', dict(old_entry) if old_entry else {}),
        old_value=dict(old_entry) if old_entry else None,
    )
    conn.execute('DELETE FROM timetable WHERE id = ?', (entry_id,))
    conn.commit()
    return jsonify({'ok': True, 'message': 'Timetable entry deleted successfully.'})


@timetable_bp.app_errorhandler(404)
def handle_404(e):
    """Custom 404 handler for reliable iframe detection."""
    return '<html><head><title>ERROR_CODE:404</title></head><body><h1>404 Not Found</h1></body></html>', 404


@timetable_bp.app_errorhandler(403)
def handle_403(e):
    """Custom 403 handler for reliable iframe detection."""
    return '<html><head><title>ERROR_CODE:403</title></head><body><h1>403 Forbidden</h1></body></html>', 403


@timetable_bp.app_errorhandler(500)
def handle_500(e):
    """Custom 500 handler for reliable iframe detection."""
    return '<html><head><title>ERROR_CODE:500</title></head><body><h1>500 Internal Server Error</h1></body></html>', 500
