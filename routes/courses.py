from io import BytesIO
import re

try:
    import pandas as pd  # type: ignore
except ImportError:
    pd = None

from flask import (
    Blueprint,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

import db
from routes import courses_timetable_admin_required, current_department_name, current_role, current_user_id, login_required
from routes.access import admin_department_required, department_name_allowed, is_super_admin

courses_bp = Blueprint('courses', __name__)

COURSE_CODE_PATTERN = re.compile(r'^[\u0647\u062a\u0639]\d{3}$')
CORRUPTED_EXCEL_TEXT_PATTERN = re.compile(r'\?{2,}|�')
HEADER_ALIASES = {
    'Course Code': {'course code', 'code', 'course_code', 'رمز المادة', 'الرمز', 'كود المادة'},
    'Course Name': {'course name', 'name', 'course_name', 'اسم المادة', 'المادة', 'اسم المقرر'},
    'Department': {'department', 'dept', 'القسم', 'التخصص'},
    'Year': {'year', 'academic year', 'السنة', 'السنة الدراسية'},
    'Notes': {'notes', 'note', 'ملاحظات', 'ملاحظة'},
}
CATEGORY_HINTS = {
    'ه': ('قسم النفط', 'قسم المدني', 'قسم المعماري', 'هندسة', 'engineering', 'engineer', 'مدني', 'كهرب', 'ميكاني', 'عمارة'),
    'ت': ('قسم الحاسوب', 'قسم الاتصالات', 'تقنية', 'technology', 'tech', 'حاسب', 'معلومات', 'برمج', 'اتصالات'),
}


def _get_departments(conn):
    return conn.execute('SELECT name FROM departments ORDER BY name').fetchall()


def _get_department_map(departments):
    return {row['name'].lower(): row['name'] for row in departments}


def _history_actor():
    return session.get('user_id'), session.get('username') or session.get('label') or 'System'


def _course_history_message(action, course):
    actor = session.get('username') or 'System'
    name = course.get('name') or f"Course #{course.get('id')}"
    department = course.get('department') or ''
    year = course.get('year') or ''
    action_word = {'ADD': 'added', 'EDIT': 'updated', 'DELETE': 'deleted'}.get(action, 'changed')
    return f"{actor} {action_word} course {name} in {department}, year {year}."


def _safe_redirect_to_courses():
    next_url = (request.args.get('next') or request.form.get('next') or '').strip()
    if next_url.startswith('/'):
        return redirect(next_url)
    return redirect(url_for('courses.list_courses'))


def _valid_department_names(departments):
    return {row['name'] for row in departments}


def _normalize_arabic_text(text):
    if text is None:
        return ""
    # Standardize 'هـ' and 'ه' + tatweel to 'ه'
    return str(text).replace('\u0647\u0640', '\u0647').replace('هـ', 'ه').strip()


def _normalize_course_code(code):
    # Unify Arabic characters and remove all internal spaces for codes
    return _normalize_arabic_text(code).replace(' ', '')


def _normalize_notes(notes):
    value = _normalize_arabic_text(notes)
    return value or None


def _parse_year(value):
    try:
        year = int(str(value).strip())
    except (TypeError, ValueError):
        return None
    return year if year in {1, 2, 3, 4} else None


def _contains_corrupted_text(value):
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    if CORRUPTED_EXCEL_TEXT_PATTERN.search(text):
        return True
    if '????' in text:
        return True
    return False


def _sanitize_excel_text(value):
    if value is None:
        return ''
    sanitized = str(value).strip()
    if _contains_corrupted_text(sanitized):
        return None
    # Normalize Arabic characters and collapse multiple spaces
    return re.sub(r'\s+', ' ', _normalize_arabic_text(sanitized))


def _infer_category_letter(department_name):
    normalized = _normalize_arabic_text(department_name).lower()
    if normalized in {'قسم النفط', 'قسم المدني', 'قسم المعماري'}:
        return 'ه'
    if normalized in {'قسم الحاسوب', 'قسم الاتصالات'}:
        return 'ت'
    if normalized == 'القسم العام':
        return 'ع'
    for letter, hints in CATEGORY_HINTS.items():
        if any(hint in normalized for hint in hints):
            return letter
    return 'ع'


def _build_generated_code(conn, department, year, exclude_course_id=None):
    letter = _infer_category_letter(department)
    prefix = f'{letter}{year}'
    
    # جلب كافة الرموز لمعالجتها في بايثون لضمان دقة التحقق من الحروف العربية
    sql = 'SELECT code FROM courses'
    params = []
    if exclude_course_id is not None:
        sql += ' WHERE id != ?'
        params.append(exclude_course_id)

    max_sequence = 0
    rows = conn.execute(sql, params).fetchall()
    for row in rows:
        norm_code = _normalize_course_code(row['code'])
        if COURSE_CODE_PATTERN.match(norm_code) and norm_code.startswith(prefix):
            try:
                # استخراج الجزء الرقمي الأخير (مثل 01 من ه101)
                seq_part = norm_code[-2:]
                max_sequence = max(max_sequence, int(seq_part))
            except (ValueError, IndexError):
                continue

    next_sequence = max_sequence + 1
    if next_sequence > 99:
        next_sequence = 99
    return f'{prefix}{next_sequence:02d}'


def _code_exists(conn, code, department, exclude_course_id=None):
    sql = 'SELECT id FROM courses WHERE department = ? AND code = ?'
    params = [department, code]
    if exclude_course_id is not None:
        sql += ' AND id != ?'
        params.append(exclude_course_id)
    return conn.execute(sql, params).fetchone() is not None


def _get_course_form_data():
    return {
        'name': _normalize_arabic_text(request.form.get('name')),
        'code': _normalize_course_code(request.form.get('code')),
        'department': _normalize_arabic_text(request.form.get('department')),
        'year': (request.form.get('year') or '').strip(),
        'notes': _normalize_arabic_text(request.form.get('notes')),
    }


def _validate_course_form(conn, form_data, departments, owner_user_id, exclude_course_id=None):
    errors = []
    dept_map = _get_department_map(departments)

    if not form_data['name']:
        errors.append('اسم المادة مطلوب.')

    dept_input = form_data['department'].lower() if form_data['department'] else ""
    if dept_input not in dept_map:
        errors.append('يرجى اختيار قسم صحيح من القائمة.')
    else:
        form_data['department'] = dept_map[dept_input]

    year = _parse_year(form_data['year'])
    if year is None:
        errors.append('السنة الدراسية يجب أن تكون بين 1 و4.')

    code = form_data['code']
    if not code:
        errors.append('رمز المادة مطلوب.')
    elif not COURSE_CODE_PATTERN.match(code):
        errors.append('الرمز يجب أن يبدأ بحرف القسم (هـ، ت، أو ع) متبوعاً بثلاثة أرقام مثل ه101 أو ت205.')
    elif year is not None and code[1] != str(year):
        errors.append('الرقم الأول بعد حرف القسم يجب أن يطابق السنة الدراسية المختارة.')
    elif _code_exists(conn, code, form_data['department'], exclude_course_id=exclude_course_id):
        errors.append('رمز المادة مستخدم بالفعل لمادة أخرى.')

    return errors, year, _normalize_notes(form_data['notes'])


def _course_payload_from_row(row):
    return {
        'name': row['name'],
        'code': _normalize_course_code(row['code']),
        'department': row['department'],
        'year': row['year'],
        'notes': row['notes'] or '',
    }


def _read_excel_rows(file_storage):
    file_bytes = file_storage.read()
    rows = []

    if pd is not None:
        try:
            df = pd.read_excel(BytesIO(file_bytes), engine='openpyxl', dtype=str, keep_default_na=False)
            raw_headers = [str(col).strip() if col is not None else '' for col in df.columns]
            values_iter = df.itertuples(index=False, name=None)
        except Exception:
            raw_headers = []
            values_iter = []
    else:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        raw_headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]] if rows else []
        values_iter = (tuple(r) for r in rows[1:])

    if not raw_headers:
        return [], ['ملف Excel فارغ.']

    header_map = {}
    for index, header in enumerate(raw_headers):
        normalized = header.lower()
        for canonical_name, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                header_map[canonical_name] = index
                break

    required_headers = ['Course Code', 'Course Name', 'Department', 'Year']
    missing_headers = [header for header in required_headers if header not in header_map]
    if missing_headers:
        return [], [f"الملف ينقصه الأعمدة المطلوبة: {', '.join(missing_headers)}."]

    parsed_rows = []
    for row_number, values in enumerate(values_iter, start=2):
        values = list(values)
        if not any(value not in (None, '') for value in values):
            continue

        raw_code = values[header_map['Course Code']] if header_map['Course Code'] < len(values) else ''
        raw_name = values[header_map['Course Name']] if header_map['Course Name'] < len(values) else ''
        raw_department = values[header_map['Department']] if header_map['Department'] < len(values) else ''
        raw_year = values[header_map['Year']] if header_map['Year'] < len(values) else ''
        raw_notes = values[header_map['Notes']] if 'Notes' in header_map and header_map['Notes'] < len(values) else ''

        parsed_rows.append(
            {
                'row_number': row_number,
                'code': _normalize_course_code(_sanitize_excel_text(raw_code) or ''),
                'name': _sanitize_excel_text(raw_name),
                'department': _sanitize_excel_text(raw_department),
                'year': _sanitize_excel_text(raw_year),
                'notes': _sanitize_excel_text(raw_notes),
                'raw_code': raw_code,
                'raw_name': raw_name,
                'raw_department': raw_department,
                'raw_notes': raw_notes,
            }
        )

    if not parsed_rows:
        return [], ['ملف Excel لا يحتوي على صفوف بيانات.']

    return parsed_rows, []


def _validate_import_rows(conn, departments, rows):
    errors = []
    dept_map = _get_department_map(departments)
    seen_codes = set()
    seen_names = set()  # (department, name_lower)
    sanitized_rows = []

    for row in rows:
        corrupted = False
        if _contains_corrupted_text(row.get('raw_code')):
            errors.append(f"الصف {row['row_number']}: رمز المادة يحتوي على قيمة غير صالحة.")
            corrupted = True
        if _contains_corrupted_text(row.get('raw_name')):
            errors.append(f"الصف {row['row_number']}: اسم المادة يحتوي على قيمة غير صالحة.")
            corrupted = True
        if _contains_corrupted_text(row.get('raw_department')):
            errors.append(f"الصف {row['row_number']}: القسم يحتوي على قيمة غير صالحة.")
            corrupted = True
        if _contains_corrupted_text(row.get('raw_notes')):
            errors.append(f"الصف {row['row_number']}: الملاحظات تحتوي على قيمة غير صالحة.")
            corrupted = True

        code = row['code']
        year = _parse_year(row['year'])
        if not corrupted and not row['name']:
            errors.append(f"الصف {row['row_number']}: اسم المادة مطلوب.")

        dept_input = row['department'].lower() if row['department'] else ""
        if not corrupted and dept_input not in dept_map:
            errors.append(f"الصف {row['row_number']}: القسم غير موجود.")
        elif not corrupted:
            row['department'] = dept_map[dept_input]

        if not corrupted and year is None:
            errors.append(f"الصف {row['row_number']}: السنة الدراسية يجب أن تكون بين 1 و4.")
        if not corrupted and not code:
            errors.append(f"الصف {row['row_number']}: رمز المادة مطلوب.")
        elif not corrupted and not COURSE_CODE_PATTERN.match(code):
            errors.append(f"الصف {row['row_number']}: الرموز المقبولة هي ه101 أو ت205 أو ع102؛ حرف القسم ثم ثلاثة أرقام.")
        elif not corrupted and year is not None and code[1] != str(year):
            errors.append(f"الصف {row['row_number']}: الرقم الأول بعد حرف القسم يجب أن يطابق السنة الدراسية المحددة.")

        name_lower = row['name'].lower() if row['name'] else ""
        dept_name_key = (row['department'], name_lower)

        if not corrupted:
            if code in seen_codes:
                errors.append(f"الصف {row['row_number']}: يوجد تكرار داخل الملف لنفس رمز المادة.")
            
            if row['name'] and dept_name_key in seen_names:
                errors.append(f"الصف {row['row_number']}: اسم المادة '{row['name']}' مكرر داخل الملف لنفس القسم.")
            
            if row['name'] and row['department'] and code:
                # Verify if this name exists in the DB for this department under a different code
                existing_name = conn.execute(
                    'SELECT code FROM courses WHERE department = ? AND LOWER(name) = ? AND code != ?',
                    (row['department'], name_lower, code)
                ).fetchone()
                if existing_name:
                    errors.append(f"الصف {row['row_number']}: المادة '{row['name']}' مسجلة مسبقاً في قسم {row['department']} برمز مختلف ({existing_name['code']}).")

        seen_codes.add(code)
        seen_names.add(dept_name_key)

        sanitized_rows.append(
            {
                'row_number': row['row_number'],
                'code': code,
                'name': row['name'],
                'department': row['department'],
                'year': year,
                'notes': _normalize_notes(row['notes']),
            }
        )

    return sanitized_rows, errors


def _upsert_course(conn, course_data):
    existing = conn.execute(
        'SELECT * FROM courses WHERE department = ? AND code = ?',
        (course_data['department'], course_data['code']),
    ).fetchone()

    if existing:
        conn.execute(
            '''
            UPDATE courses
            SET name = ?, department = ?, year = ?, notes = ?
            WHERE id = ?
            ''',
            (
                course_data['name'],
                course_data['department'],
                course_data['year'],
                course_data['notes'],
                existing['id'],
            ),
        )
        updated = conn.execute('SELECT * FROM courses WHERE id = ?', (existing['id'],)).fetchone()
        db.add_history(
            conn,
            'EDIT',
            'course',
            existing['id'],
            *_history_actor(),
            message=_course_history_message('EDIT', dict(updated) if updated else dict(existing)),
            old_value=dict(existing),
            new_value=dict(updated) if updated else None,
        )
        return 'updated'

    cursor = conn.execute(
        '''
        INSERT INTO courses (user_id, name, code, department, year, notes)
        VALUES (?, ?, ?, ?, ?, ?)
        ''',
        (
            current_user_id(),
            course_data['name'],
            course_data['code'],
            course_data['department'],
            course_data['year'],
            course_data['notes'],
        ),
    )
    inserted = conn.execute('SELECT * FROM courses WHERE id = ?', (cursor.lastrowid,)).fetchone()
    db.add_history(
        conn,
        'ADD',
        'course',
        cursor.lastrowid,
        *_history_actor(),
        message=_course_history_message('ADD', dict(inserted) if inserted else course_data),
        new_value=dict(inserted) if inserted else course_data,
    )
    return 'inserted'



@courses_bp.route('/')
@login_required
@admin_department_required
def list_courses():
    conn = db.get_db()
    requested_department = (request.args.get('department') or '').strip()
    department = requested_department if is_super_admin() else (current_department_name(conn) or '')
    search = (request.args.get('q') or '').strip()
    selected_year = _parse_year(request.args.get('year'))
    departments = _get_departments(conn) if is_super_admin() else [{'name': current_department_name(conn)}]

    if is_super_admin():
        sql = '''
        SELECT c.*, u.label AS owner_label
        FROM courses c
        LEFT JOIN users u ON u.id = c.user_id
        WHERE 1=1
        '''
        params = []
        if department:
            sql += ' AND c.department = ?'
            params.append(department)
        if selected_year is not None:
            sql += ' AND c.year = ?'
            params.append(selected_year)
        if search:
            sql += ' AND (c.name LIKE ? OR c.code LIKE ?)'
            params.extend([f'%{search}%', f'%{search}%'])
        sql += ' ORDER BY c.year, c.code, c.name'
        courses = conn.execute(sql, params).fetchall()
    else:
        # الأدمن يرى فقط دورات قسمه
        sql = '''
        SELECT
            c.*,
            GROUP_CONCAT(DISTINCT te.name) AS teachers,
            GROUP_CONCAT(DISTINCT t.semester) AS semesters
        FROM courses c
        LEFT JOIN timetable t ON c.id = t.course_id AND t.user_id = c.user_id
        LEFT JOIN teachers te ON t.teacher_id = te.id
        WHERE c.department = ?
        '''
        params = [current_department_name(conn)]
        if selected_year is not None:
            sql += ' AND c.year = ?'
            params.append(selected_year)
        if search:
            sql += ' AND (c.name LIKE ? OR c.code LIKE ?)'
            params.extend([f'%{search}%', f'%{search}%'])
        sql += ' GROUP BY c.id ORDER BY c.year, c.code, c.name'
        courses = conn.execute(sql, params).fetchall()

    return render_template(
        'courses/list_clean.html',
        courses=courses,
        department=department,
        departments=departments,
        search=search,
        selected_year=selected_year or '',
    )


@courses_bp.route('/create', methods=['GET', 'POST'])
@login_required
@courses_timetable_admin_required
def create_course():
    conn = db.get_db()
    departments = _get_departments(conn) if is_super_admin() else [{'name': current_department_name(conn)}]
    form_data = {'name': '', 'code': '', 'department': '', 'year': '1', 'notes': ''}

    if request.method == 'POST':
        form_data = _get_course_form_data()
        if not is_super_admin():
            form_data['department'] = current_department_name(conn) or ''
        errors, year, notes = _validate_course_form(conn, form_data, departments, owner_user_id=current_user_id())

        for error in errors:
            flash(error, 'danger')

        if not errors:
            cursor = conn.execute(
                '''
                INSERT INTO courses (user_id, name, code, department, year, notes)
                VALUES (?, ?, ?, ?, ?, ?)
                ''',
                (
                    current_user_id(),
                    form_data['name'],
                    form_data['code'],
                    form_data['department'],
                    year,
                    notes,
                ),
            )
            inserted = conn.execute('SELECT * FROM courses WHERE id = ?', (cursor.lastrowid,)).fetchone()
            db.add_history(
                conn,
                'ADD',
                'course',
                cursor.lastrowid,
                *_history_actor(),
                message=_course_history_message('ADD', dict(inserted) if inserted else {'name': form_data['name'], 'code': form_data['code'], 'department': form_data['department'], 'year': year}),
                new_value=dict(inserted) if inserted else {'name': form_data['name'], 'code': form_data['code'], 'department': form_data['department'], 'year': year},
            )
            conn.commit()
            flash('تم إضافة المقرر بنجاح.', 'success')
            return _safe_redirect_to_courses()

    return render_template('courses/create.html', departments=departments, form_data=form_data)


@courses_bp.route('/upload', methods=['POST'])
@login_required
@courses_timetable_admin_required
def upload_courses_excel():
    conn = db.get_db()
    departments = _get_departments(conn) if is_super_admin() else [{'name': current_department_name(conn)}]
    upload = request.files.get('excel_file')

    if not upload or not upload.filename:
        flash('يرجى اختيار ملف Excel للرفع.', 'danger')
        return _safe_redirect_to_courses()

    if not upload.filename.lower().endswith('.xlsx'):
        flash('صيغة الملف غير مدعومة. يرجى رفع ملف .xlsx فقط.', 'danger')
        return _safe_redirect_to_courses()

    try:
        rows, file_errors = _read_excel_rows(upload)
    except Exception:
        flash('تعذر قراءة ملف Excel. تأكد من أن الملف بصيغة .xlsx وصالح للقراءة.', 'danger')
        return _safe_redirect_to_courses()
    sanitized_rows, validation_errors = _validate_import_rows(conn, departments, rows)
    if not is_super_admin():
        current_department = current_department_name(conn) or ''
        for row in sanitized_rows:
            row['department'] = current_department
    errors = file_errors + validation_errors

    if errors:
        for error in errors[:8]:
            flash(error, 'danger')
        if len(errors) > 8:
            flash(f'يوجد {len(errors) - 8} أخطاء إضافية في الملف.', 'danger')
        return redirect(url_for('courses.list_courses'))

    inserted = 0
    updated = 0
    for row in sanitized_rows:
        action = _upsert_course(conn, row)
        if action == 'inserted':
            inserted += 1
        else:
            updated += 1

    conn.commit()
    flash(f'تمت معالجة ملف Excel بنجاح. تمت إضافة {inserted} مادة وتحديث {updated} مادة.', 'success')
    return redirect(url_for('courses.list_courses'))


@courses_bp.route('/download')
@login_required
@courses_timetable_admin_required
def download_courses_excel():
    conn = db.get_db()
    if is_super_admin():
        courses = conn.execute(
            '''
            SELECT code, name, department, year, notes
            FROM courses
            ORDER BY department, year, code
            '''
        ).fetchall()
    else:
        courses = conn.execute(
            '''
            SELECT code, name, department, year, notes
            FROM courses
            WHERE department = ?
            ORDER BY department, year, code
            ''',
            (current_department_name(conn),),
        ).fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Courses'
    headers = ['Course Code', 'Course Name', 'Department', 'Year', 'Notes']
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for course in courses:
        sheet.append(
            [
                _normalize_course_code(course['code']),
                course['name'],
                course['department'],
                course['year'],
                course['notes'] or '',
            ]
        )

    for column in sheet.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 4, 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='courses.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@courses_bp.route('/generate-code')
@login_required
@courses_timetable_admin_required
def generate_course_code():
    conn = db.get_db()
    department = request.args.get('department', '').strip()
    if not is_super_admin():
        department = current_department_name(conn) or ''
    year = _parse_year(request.args.get('year'))
    course_id = request.args.get('course_id', type=int)

    if not department or year is None:
        return jsonify({'error': 'department and year are required'}), 400

    return jsonify({'code': _build_generated_code(conn, department, year, exclude_course_id=course_id)})


@courses_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@courses_timetable_admin_required
def edit_course(id):
    conn = db.get_db()
    course = conn.execute('SELECT * FROM courses WHERE id = ?', (id,)).fetchone()
    departments = _get_departments(conn) if is_super_admin() else [{'name': current_department_name(conn)}]

    if not course:
        flash('المقرر غير موجود.', 'danger')
        return redirect(url_for('courses.list_courses'))
    if not is_super_admin() and not department_name_allowed(conn, course['department']):
        flash('ليس لديك صلاحية الوصول.', 'danger')
        return redirect(url_for('courses.list_courses'))

    form_data = _course_payload_from_row(course)

    if request.method == 'POST':
        form_data = _get_course_form_data()
        if not is_super_admin():
            form_data['department'] = current_department_name(conn) or ''
        errors, year, notes = _validate_course_form(
            conn,
            form_data,
            departments,
            owner_user_id=course['user_id'],
            exclude_course_id=id,
        )

        for error in errors:
            flash(error, 'danger')

        if not errors:
            conn.execute(
                '''
                UPDATE courses
                SET name = ?, code = ?, department = ?, year = ?, notes = ?
                WHERE id = ?
                ''',
                (
                    form_data['name'],
                    form_data['code'],
                    form_data['department'],
                    year,
                    notes,
                    id,
                ),
            )
            updated = conn.execute('SELECT * FROM courses WHERE id = ?', (id,)).fetchone()
            db.add_history(
                conn,
                'EDIT',
                'course',
                id,
                *_history_actor(),
                message=_course_history_message('EDIT', dict(updated) if updated else {'name': form_data['name'], 'code': form_data['code'], 'department': form_data['department'], 'year': year}),
                old_value=dict(course),
                new_value=dict(updated) if updated else None,
            )
            conn.commit()
            flash('تم تحديث المقرر بنجاح.', 'success')
            return _safe_redirect_to_courses()

    return render_template(
        'courses/edit.html',
        course=course,
        departments=departments,
        form_data=form_data,
    )


@courses_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
@courses_timetable_admin_required
def delete_course(id):
    conn = db.get_db()
    course = conn.execute('SELECT * FROM courses WHERE id = ?', (id,)).fetchone()
    if not course:
        flash('المقرر غير موجود.', 'danger')
        return _safe_redirect_to_courses()
    if not is_super_admin() and not department_name_allowed(conn, course['department']):
        flash('ليس لديك صلاحية حذف مقرر من قسم آخر.', 'danger')
        return _safe_redirect_to_courses()

    db.add_history(
        conn,
        'DELETE',
        'course',
        id,
        *_history_actor(),
        message=_course_history_message('DELETE', dict(course)),
        old_value=dict(course),
    )
    conn.execute('DELETE FROM courses WHERE id = ?', (id,))
    conn.commit()
    flash('تم حذف المقرر بنجاح.', 'success')
    return _safe_redirect_to_courses()
