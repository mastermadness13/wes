from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
import db

from routes import current_department_id, current_role, login_required
from routes.access import admin_department_required, is_super_admin

rooms_bp = Blueprint('rooms', __name__)

@rooms_bp.route('/')
@login_required
@admin_department_required
def list_rooms():
    conn = db.get_db()
    search_query = request.args.get('q', '')
    room_type = request.args.get('type', '')
    status = request.args.get('status', '')
    location = request.args.get('location', '')

    query = 'SELECT * FROM rooms WHERE 1=1'
    params = []

    if not is_super_admin():
        query += ' AND department_id = ?'
        params.append(current_department_id())

    if search_query:
        query += ' AND (name LIKE ? OR name_ar LIKE ?)'
        params.extend([f'%{search_query}%', f'%{search_query}%'])

    if room_type:
        query += ' AND type = ?'
        params.append(room_type)
    if status:
        query += ' AND status = ?'
        params.append(status)
    if location:
        query += ' AND location = ?'
        params.append(location)

    query += ' ORDER BY name'
    rooms = conn.execute(query, params).fetchall()

    return render_template('rooms/list.html', rooms=rooms, room_type=room_type, status=status, location=location, q=search_query)

@rooms_bp.route('/export')
@login_required
@admin_department_required
def export_rooms():
    conn = db.get_db()
    search_query = request.args.get('q', '')
    room_type = request.args.get('type', '')
    status = request.args.get('status', '')
    location = request.args.get('location', '')

    query = 'SELECT * FROM rooms WHERE 1=1'
    params = []

    if not is_super_admin():
        query += ' AND department_id = ?'
        params.append(current_department_id())

    if search_query:
        query += ' AND (name LIKE ? OR name_ar LIKE ?)'
        params.extend([f'%{search_query}%', f'%{search_query}%'])

    if room_type:
        query += ' AND type = ?'
        params.append(room_type)
    if status:
        query += ' AND status = ?'
        params.append(status)
    if location:
        query += ' AND location = ?'
        params.append(location)

    query += ' ORDER BY name'
    rooms = conn.execute(query, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rooms"
    
    # ضبط اتجاه الصفحة من اليمين لليسار (للغة العربية)
    ws.sheet_view.rightToLeft = True

    # 1. إضافة شعار المؤسسة (يجب التأكد من وجود ملف logo.png في مجلد static/img)
    try:
        img = Image('static/img/logo.png')
        img.height = 60
        img.width = 60
        ws.add_image(img, 'A1')
    except:
        pass  # في حال عدم وجود الصورة، سيتم تخطي هذه الخطوة

    # 2. إضافة عنوان التقرير وتنسيقه
    ws.merge_cells('B1:F2')
    report_title = ws['B1']
    report_title.value = "تقرير بيانات القاعات والأقسام"
    report_title.font = Font(size=16, bold=True, color="1457d2")
    report_title.alignment = Alignment(horizontal='center', vertical='center')

    # 3. تعريف تنسيق العناوين (Header Styling)
    header_fill = PatternFill(start_color='1457d2', end_color='1457d2', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 4. كتابة العناوين في الصف الرابع (لترك مساحة للشعار)
    start_row = 4
    headers = ['اسم القاعة', 'النوع', 'الحالة', 'المقاعد', 'الأجهزة', 'الموقع']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # 5. تعبئة البيانات
    for row_num, r in enumerate(rooms, start_row + 1):
        ws.cell(row=row_num, column=1, value=r['name_ar'] or r['name'])
        ws.cell(row=row_num, column=2, value=r['type'])
        ws.cell(row=row_num, column=3, value=r['status'])
        ws.cell(row=row_num, column=4, value=r['capacity'])
        ws.cell(row=row_num, column=5, value=r['devices'])
        ws.cell(row=row_num, column=6, value=r['location'])
        
        # محاذاة البيانات للوسط
        for col_num in range(1, 7):
            ws.cell(row=row_num, column=col_num).alignment = center_alignment

    # 6. ضبط عرض الأعمدة تلقائياً
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) # إضافة مسافة بادئة صغيرة
        ws.column_dimensions[column_letter].width = adjusted_width

    # 7. إضافة إطارات (Borders) لكافة خلايا الجدول
    thin_side = Side(border_style="thin", color="000000")
    table_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = table_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='rooms_export.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@rooms_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_room():
    if request.method == 'POST':
        name = request.form.get('name')
        name_ar = request.form.get('name_ar')
        # إذا لم يُرسل name استخدم name_ar
        if not name and name_ar:
            name = name_ar
        type_ = request.form.get('type')
        status = request.form.get('status')
        capacity = request.form.get('capacity', type=int)
        devices = request.form.get('devices', type=int)
        location = request.form.get('location')

        if not name:
            flash('يجب إدخال اسم القاعة بأي لغة', 'danger')
            return redirect(request.referrer or url_for('rooms.list_rooms'))

        conn = db.get_db()
        try:
            conn.execute('INSERT INTO rooms (name, name_ar, type, status, capacity, devices, location, department_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                        (name, name_ar, type_, status, capacity, devices, location, current_department_id()))
            conn.commit()
            flash('تم إضافة القاعة بنجاح', 'success')
        except Exception as e:
            flash(f'خطأ: {str(e)}', 'danger')
        return redirect(url_for('rooms.list_rooms'))

    return render_template('rooms/create.html')

@rooms_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_room(id):
    conn = db.get_db()
    room = conn.execute('SELECT * FROM rooms WHERE id = ?', (id,)).fetchone()
    # حماية: لا يسمح إلا لصاحب الصلاحية
    if room and not is_super_admin() and room['department_id'] != current_department_id():
        flash('ليس لديك صلاحية تعديل قاعة من قسم آخر.', 'danger')
        return redirect(url_for('rooms.list_rooms'))
    if not room:
        flash('القاعة غير موجودة', 'danger')
        return redirect(url_for('rooms.list_rooms'))

    if request.method == 'POST':
        name = request.form.get('name')
        name_ar = request.form.get('name_ar')
        # إذا لم يُرسل name استخدم name_ar
        if not name and name_ar:
            name = name_ar
        type_ = request.form.get('type')
        status = request.form.get('status')
        capacity = request.form.get('capacity', type=int)
        devices = request.form.get('devices', type=int)
        location = request.form.get('location')

        if not name:
            flash('يجب إدخال اسم القاعة بأي لغة', 'danger')
            return redirect(request.referrer or url_for('rooms.list_rooms'))

        try:
            conn.execute('UPDATE rooms SET name = ?, name_ar = ?, type = ?, status = ?, capacity = ?, devices = ?, location = ? WHERE id = ?',
                        (name, name_ar, type_, status, capacity, devices, location, id))
            conn.commit()
            flash('تم تحديث القاعة بنجاح', 'success')
        except Exception as e:
            flash(f'خطأ: {str(e)}', 'danger')
        return redirect(url_for('rooms.list_rooms'))

    return render_template('rooms/edit.html', room=room)

@rooms_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_room(id):
    conn = db.get_db()
    room = conn.execute('SELECT * FROM rooms WHERE id = ?', (id,)).fetchone()
    if not room:
        flash('القاعة غير موجودة', 'danger')
        return redirect(url_for('rooms.list_rooms'))
    # حماية: لا يسمح إلا لصاحب الصلاحية
    if not is_super_admin() and room['department_id'] != current_department_id():
        flash('ليس لديك صلاحية حذف قاعة من قسم آخر.', 'danger')
        return redirect(url_for('rooms.list_rooms'))
    try:
        conn.execute('DELETE FROM rooms WHERE id = ?', (id,))
        conn.commit()
        flash('تم حذف القاعة بنجاح', 'success')
    except Exception as e:
        flash(f'خطأ: {str(e)}', 'danger')
    return redirect(url_for('rooms.list_rooms'))
